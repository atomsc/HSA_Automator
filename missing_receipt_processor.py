"""
Module to process missing receipts and update the Excel file.
"""

import os
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill
from browser_setup import BrowserSetup
from pdf_receipt_processor import PDFReceiptProcessor
from receipt_uploader import ReceiptUploader


class MissingReceiptProcessor:
    """
    This class processes missing receipts and updates the Excel file accordingly.
    """

    def __init__(self, df: pd.DataFrame, excel_file_loc: str, directory_path: str, url: str, email: str, password: str):
        """
        Initializes the MissingReceiptProcessor instance.

        Parameters:
        df (pd.DataFrame): DataFrame containing the data.
        excel_file_loc (str): Location of the Excel file.
        directory_path (str): Directory containing the receipts.
        URL (str): URL of the login page.
        EMAIL (str): User's email address.
        PASSWORD (str): User's password.
        """
        self.df_not_uploaded = None
        self.df = df
        self.excel_file_loc = excel_file_loc
        self.sheet_name = "TransactionHistory (2)"
        self.search_column = "Receipt no"
        self.update_column = "Attachments"
        self.new_value = "Y"
        self.directory_path = directory_path
        self.URL = url
        self.EMAIL = email
        self.PASSWORD = password
        self.browser = BrowserSetup(url, email, password)

    def filter_not_uploaded(self):
        """
        Filters the DataFrame to find receipts that have not been uploaded.
        """
        df_not_uploaded = self.df[self.df['Attachments'] != "Y"].copy()
        df_not_uploaded['Date'] = pd.to_datetime(df_not_uploaded['Date']).dt.strftime('%m/%d/%Y')
        self.df_not_uploaded = df_not_uploaded

    def display_not_uploaded(self):
        """
        Displays the receipts that have not been uploaded.
        """
        columns_to_display = ['Date', 'Provider', 'Amount', 'Receipt no']
        print("Receipts not uploaded:")
        print(self.df_not_uploaded[columns_to_display])

    def select_receipt(self) -> pd.Series:
        """
        Prompts the user to select a receipt to upload.

        Returns:
        pd.Series: Selected receipt data.
        """
        while True:
            try:
                index = int(input("Enter the index of the receipt you want to upload: "))
                if index in self.df_not_uploaded.index:
                    print(f"You selected receipt at index {index}:")
                    print(self.df_not_uploaded.loc[index])
                    return self.df_not_uploaded.loc[index]
                else:
                    print("Invalid index. Please try again.")
            except ValueError:
                print("Please enter a valid integer index.")

    def find_column_letters(self, sheet) -> tuple[str, str]:
        """
        Finds the column letters for the search and update columns in the Excel sheet.

        Parameters:
        sheet (Worksheet): Excel worksheet.

        Returns:
        tuple[str, str]: Column letters for the search and update columns.
        """
        search_col_letter = None
        update_col_letter = None
        for col in sheet.iter_cols(1, sheet.max_column):
            if col[0].value == self.search_column:
                search_col_letter = col[0].column_letter
            if col[0].value == self.update_column:
                update_col_letter = col[0].column_letter
        return search_col_letter, update_col_letter

    def update_row(self, row, search_value: str, update_col_letter: str, sheet):
        """
        Updates a row in the Excel sheet.

        Parameters:
        row (Row): Row to update.
        search_value (str): Value to search for.
        update_col_letter (str): Column letter of the update column.
        sheet (Worksheet): Excel worksheet.
        """
        row[sheet[update_col_letter + '1'].col_idx - 1].value = self.new_value
        print(f"Updated {self.update_column} in row {row[0].row} to {self.new_value}.")
        if search_value.startswith('F'):
            fill = PatternFill(start_color="00B0F0", end_color="00B0F0", fill_type="solid")
        elif search_value.startswith('R'):
            fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
        else:
            print("Error: search_value must begin with 'F' or 'R'.")
            return
        for cell in row:
            cell.fill = fill

    def insert_into_cell(self, search_value: str):
        """
        Inserts a value into the appropriate cell in the Excel sheet.

        Parameters:
        search_value (str): Value to insert.
        """
        wb = openpyxl.load_workbook(self.excel_file_loc)
        sheet = wb[self.sheet_name]
        search_col_letter, update_col_letter = self.find_column_letters(sheet)
        if not search_col_letter or not update_col_letter:
            print(f"Error: Column '{self.search_column}' or '{self.update_column}' not found.")
            return
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
            if row[sheet[search_col_letter + '1'].col_idx - 1].value == search_value:
                self.update_row(row, search_value, update_col_letter, sheet)
                break
        else:
            print(f"Value '{search_value}' not found in column '{self.search_column}'.")
        wb.save(self.excel_file_loc)
        print(f"Saved changes to {self.excel_file_loc}.")

    def process_receipt_selection(self):
        """
        Processes the selection and uploading of missing receipts.
        """
        self.filter_not_uploaded()
        self.display_not_uploaded()
        selected_receipt = self.select_receipt().copy()
        name_to_rename = selected_receipt['Receipt no']
        pdf_files = [f for f in os.listdir(self.directory_path) if f.lower().endswith('.pdf')]
        for pdf_receipt in pdf_files:
            pdf_renamer = PDFReceiptProcessor(self.directory_path, self.excel_file_loc)
            file_path = os.path.join(self.directory_path, pdf_receipt)
            pdf_renamer.display_pdf_info(file_path, pdf_receipt)
            is_correct_receipt = input("Is this the receipt you want to match? Y or N ")
            if is_correct_receipt == "Y":
                print("YOU FOUND IT")
                new_file_path = os.path.join(self.directory_path, name_to_rename + f"_{selected_receipt['Amount']}.pdf")
                os.rename(file_path, new_file_path)
                selected_receipt['In HSA?'] = "Y"
                selected_receipt['New Filename'] = f"{name_to_rename}_{selected_receipt['Amount']}"
                uploader = ReceiptUploader(self.browser, selected_receipt.to_frame().T, self.directory_path,
                                           os.path.join(self.directory_path, "Receipts"))
                uploader.search_and_upload_receipt()
                self.insert_into_cell(name_to_rename)
                break
