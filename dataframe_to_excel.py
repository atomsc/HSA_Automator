"""
Module to handle processing and saving DataFrames to Excel.
"""

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


class DataFrameToExcel:
    """
    This class handles the processing and saving of DataFrames to Excel.
    """

    def __init__(self, df: pd.DataFrame, workbook_path: str):
        """
        Initializes the DataFrameToExcel instance.

        Parameters:
        df (pd.DataFrame): DataFrame containing the data.
        workbook_path (str): Path to the Excel workbook.
        """
        self.df = df
        self.workbook_path = workbook_path
        self.workbook = load_workbook(self.workbook_path)
        self.sheet = self.workbook.active
        self.fill_colors = {
            'R': PatternFill(start_color="92D050", end_color="92D050", fill_type="solid"),
            'F': PatternFill(start_color="00B0F0", end_color="00B0F0", fill_type="solid"),
            'default': PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        }
        self.thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

    def preprocess_data(self):
        """
        Preprocesses the data in the DataFrame.
        """
        self.df['Amount'] = pd.to_numeric(self.df['Amount'], errors='coerce')
        self.df['Date'] = pd.to_datetime(self.df['Date'], format='%m/%d/%Y')
        self.df.sort_values(by='Date', ascending=False, inplace=True)

    def apply_cell_styles(self, cell, col_name: str):
        """
        Applies styles to a cell in the Excel sheet.

        Parameters:
        cell (Cell): Cell to apply styles to.
        col_name (str): Name of the column the cell belongs to.
        """
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = self.thin_border
        if col_name.lower() == 'date':
            cell.number_format = 'MM/DD/YYYY'
        elif col_name.lower() == 'amount':
            cell.number_format = '0.00'

    def get_fill_color(self, receipt_number: str) -> PatternFill:
        """
        Gets the fill color based on the receipt number prefix.

        Parameters:
        receipt_number (str): Receipt number.

        Returns:
        PatternFill: Fill color for the cell.
        """
        prefix = receipt_number[0] if isinstance(receipt_number, str) else ''
        return self.fill_colors.get(prefix, self.fill_colors['default'])

    def insert_data_into_sheet(self):
        """
        Inserts data from the DataFrame into the Excel sheet.
        """
        for index, row in self.df.iterrows():
            self.sheet.insert_rows(2)
            receipt_number = str(row['Receipt Number'])
            fill_color = self.get_fill_color(receipt_number)
            for col_num, (col_name, value) in enumerate(row.items(), start=1):
                cell = self.sheet.cell(row=2, column=col_num, value=value)
                self.apply_cell_styles(cell, col_name)
                cell.fill = fill_color

    def adjust_column_widths(self):
        """
        Adjusts the column widths in the Excel sheet based on content.
        """
        for col_num in range(1, len(self.df.columns) + 1):
            col_letter = get_column_letter(col_num)
            self.sheet.column_dimensions[col_letter].auto_size = True

    def save_workbook(self):
        """
        Saves the workbook with changes.
        """
        self.workbook.save(self.workbook_path)

    def process(self):
        """
        Processes the DataFrame and saves it to the Excel sheet.
        """
        self.preprocess_data()
        self.insert_data_into_sheet()
        self.adjust_column_widths()
        self.save_workbook()
